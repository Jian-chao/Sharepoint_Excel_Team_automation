"""
modules/excel_reader.py
========================
Color-and-merge-aware Excel parser for the FDI DEF request status sheet.

Key features:
  • Resolves merged cells (reads top-left value for any cell in a range)
  • Detects pre-filled (blue) vs. user-filled (yellow) columns via bg color
  • Flags rows whose B column has a non-black font color as skip=True
  • Provides generate_llm_prompt() to build a structured prompt for AI assistants

Usage:
    from modules.excel_reader import ExcelReader
    reader = ExcelReader()
    records = reader.get_all_records()   # list[SubsysRecord]
    prompt  = reader.generate_llm_prompt("Owner note: row 7 is TBD scope")
"""

import io
import logging
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX

import config

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# Data model
# ─────────────────────────────────────────────────────────────

@dataclass
class SubsysRecord:
    row:          int
    subsys:       str
    fe:           str
    bc:           str
    be:           str
    ppt_status:   str       # "done", "eta:YYYY-MM-DD", "n/a", or ""
    netlist:      str       # "v", "x", or ""
    sdc:          str
    ccf:          str
    upf:          str
    release_date: str
    remarks:      str
    skip:         bool      # True → non-black font on B col (grey-out)
    font_color_b: str       # raw RGB of col B font (useful for debugging)


# ─────────────────────────────────────────────────────────────
# Helper — cell color extraction
# ─────────────────────────────────────────────────────────────

_BLACK_RGBS = {config.COLOR_BLACK, config.COLOR_DEFAULT, "00000000", "FF000000"}

# In the default Office theme:
#   theme=1, tint=0.0  →  dk1 (dark 1) = black text  (normal/default)
#   theme=0, tint<0    →  lt1 (light 1) darkened by |tint| = grey
# We treat theme=1 with tint=0.0 as "black" and everything else as "non-black".
_THEME_BLACK = (1, 0.0)   # (theme_index, tint) that equals the default black text


def _get_rgb(color_obj) -> str:
    """
    Return a canonical color descriptor from an openpyxl Color object.
    For theme colors returns 'THEME:<index>:<tint>' so callers can inspect them.
    Falls back to '00000000' (default/transparent = treat as black).
    """
    if color_obj is None:
        return "00000000"
    try:
        if color_obj.type == "rgb":
            return color_obj.rgb or "00000000"
        if color_obj.type == "indexed":
            return f"IDX:{color_obj.indexed}"
        if color_obj.type == "theme":
            tint = getattr(color_obj, "tint", 0.0) or 0.0
            return f"THEME:{color_obj.theme}:{tint}"
    except Exception:
        pass
    return "00000000"


def _is_non_black(rgb: str) -> bool:
    """
    Return True if the color is clearly not the default black.

    Strategy:
    - RGB 'FF000000' / '00000000'  → black / default             → False
    - THEME:1:0.0                   → dk1 with no tint = black    → False
    - THEME:<other>:<tint>          → any other theme color        → True
    - IDX:<n>                       → indexed color                → True (assume coloured)
    """
    if rgb in _BLACK_RGBS:
        return False
    if rgb.startswith("00"):
        # Fully transparent = Excel default = treat as black
        return False
    if rgb.startswith("THEME:"):
        # Parse 'THEME:<index>:<tint>'
        parts = rgb.split(":")
        try:
            theme_idx = int(parts[1])
            tint      = float(parts[2]) if len(parts) > 2 else 0.0
            return not (theme_idx == _THEME_BLACK[0] and abs(tint) < 1e-6)
        except (ValueError, IndexError):
            return True  # unrecognisable theme — treat as coloured
    if rgb.startswith("IDX:"):
        return True  # indexed colour
    return True  # explicit non-black RGB


# ─────────────────────────────────────────────────────────────
# Main reader
# ─────────────────────────────────────────────────────────────

class ExcelReader:
    """
    Reads the FDI DEF request status sheet.

    Parameters
    ----------
    excel_bytes : bytes, optional
        Raw bytes of the workbook.  If None, reads from LOCAL_EXCEL_PATH.
    """

    def __init__(self, excel_bytes: Optional[bytes] = None):
        if excel_bytes is None:
            with open(config.LOCAL_EXCEL_PATH, "rb") as fh:
                excel_bytes = fh.read()

        self._wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        self._ws = self._wb[config.EXCEL_SHEET_NAME]
        self._merge_map = self._build_merge_map()

    # ----------------------------------------------------------
    # Merge-cell resolution
    # ----------------------------------------------------------
    def _build_merge_map(self) -> dict:
        """
        Build a mapping from every merged cell → its top-left (master) cell coordinate.
        e.g. "H1" → "G1"  (because G1:J1 is merged)
        """
        merge_map = {}
        for merged_range in self._ws.merged_cells.ranges:
            master = (merged_range.min_row, merged_range.min_col)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != master:
                        merge_map[(row, col)] = master
        return merge_map

    def _cell_value(self, row: int, col: int):
        """Return the effective value of a cell, resolving merged ranges."""
        master = self._merge_map.get((row, col))
        if master:
            row, col = master
        v = self._ws.cell(row=row, column=col).value
        return "" if v is None else str(v).strip()

    def _cell_obj(self, row: int, col: int):
        """Return the actual openpyxl cell object (un-merged address)."""
        master = self._merge_map.get((row, col))
        if master:
            row, col = master
        return self._ws.cell(row=row, column=col)

    # ----------------------------------------------------------
    # Public API
    # ----------------------------------------------------------
    def get_all_records(self, include_skipped: bool = False) -> list[SubsysRecord]:
        """
        Parse every data row and return a list of SubsysRecord.

        Parameters
        ----------
        include_skipped : bool
            If False (default), rows flagged skip=True are excluded.
        """
        records = []
        ws = self._ws

        for row_idx in range(config.EXCEL_DATA_START_ROW, ws.max_row + 1):
            subsys = self._cell_value(row_idx, config.COL_SUBSYS)
            if not subsys:
                continue   # blank row — stop

            # Skip-flag: any non-black font on the B column
            b_cell      = self._cell_obj(row_idx, config.COL_SUBSYS)
            font_rgb    = _get_rgb(b_cell.font.color) if b_cell.font and b_cell.font.color else "00000000"
            skip_row    = _is_non_black(font_rgb)

            rec = SubsysRecord(
                row          = row_idx,
                subsys       = subsys,
                fe           = self._cell_value(row_idx, config.COL_FE),
                bc           = self._cell_value(row_idx, config.COL_BC),
                be           = self._cell_value(row_idx, config.COL_BE),
                ppt_status   = self._cell_value(row_idx, config.COL_PPT),
                netlist      = self._cell_value(row_idx, config.COL_NETLIST),
                sdc          = self._cell_value(row_idx, config.COL_SDC),
                ccf          = self._cell_value(row_idx, config.COL_CCF),
                upf          = self._cell_value(row_idx, config.COL_UPF),
                release_date = self._cell_value(row_idx, config.COL_RELEASE_DATE),
                remarks      = self._cell_value(row_idx, config.COL_REMARKS),
                skip         = skip_row,
                font_color_b = font_rgb,
            )
            if include_skipped or not skip_row:
                records.append(rec)

        logger.info(f"[ExcelReader] Parsed {len(records)} record(s) from sheet '{config.EXCEL_SHEET_NAME}'")
        return records

    def get_summary_dict(self) -> dict:
        """
        Returns a nested dict keyed by subsys name for quick lookups.
        {subsys_name: SubsysRecord, ...}
        """
        return {r.subsys: r for r in self.get_all_records()}

    # ----------------------------------------------------------
    # LLM Prompt Generator
    # ----------------------------------------------------------
    def generate_llm_prompt(self, owner_notes: str = "") -> str:
        """
        Scan the sheet structure and produce a detailed textual description
        suitable for pasting into ChatGPT / Gemini to help an AI assistant
        understand how to parse or interpret this spreadsheet.

        Parameters
        ----------
        owner_notes : str
            Free-text annotations from the program owner, appended at the end.

        Returns
        -------
        str
            A formatted prompt string.
        """
        ws = self._ws
        lines = []
        lines.append("=" * 70)
        lines.append("EXCEL SHEET DESCRIPTION FOR LLM-ASSISTED PARSING")
        lines.append("=" * 70)
        lines.append(f"Workbook : {config.LOCAL_EXCEL_PATH}")
        lines.append(f"Sheet    : {config.EXCEL_SHEET_NAME}")
        lines.append(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        lines.append("")

        # Merged cells
        lines.append("── Merged Cell Ranges ─────────────────────────────────────────────")
        for mr in ws.merged_cells.ranges:
            master_cell = ws.cell(mr.min_row, mr.min_col)
            lines.append(f"  {mr}  →  master value = {repr(master_cell.value)}")
        lines.append("")

        # Header rows (rows 1-3)
        lines.append("── Row-by-Row Header Analysis (rows 1–3) ─────────────────────────")
        for r in range(1, config.EXCEL_DATA_START_ROW):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                val  = cell.value
                if val is None and (r, c) not in self._merge_map:
                    continue
                bg_rgb = _get_rgb(cell.fill.fgColor) if cell.fill else "00000000"
                fc_rgb = _get_rgb(cell.font.color)   if cell.font else "00000000"
                label  = _color_name(bg_rgb)
                lines.append(
                    f"  {cell.coordinate:5s}: {repr(str(val)):<45s} "
                    f"bg={bg_rgb} ({label})  font={fc_rgb}"
                )
        lines.append("")

        # Data rows
        lines.append("── Data Rows ───────────────────────────────────────────────────────")
        all_records = self.get_all_records(include_skipped=True)
        for rec in all_records:
            skip_tag = " ⚠ SKIP (non-black font)" if rec.skip else ""
            lines.append(
                f"  Row {rec.row}: subsys={rec.subsys!r:30s}  FE={rec.fe!r}  "
                f"BC={rec.bc!r}  BE={rec.be!r}{skip_tag}"
            )
            lines.append(
                f"         ppt={rec.ppt_status!r}  netlist={rec.netlist!r}  "
                f"sdc={rec.sdc!r}  ccf={rec.ccf!r}  upf={rec.upf!r}"
            )
        lines.append("")

        # Column role mapping
        lines.append("── Column Role Mapping ─────────────────────────────────────────────")
        col_role = {
            "A": ("Pushdown flag", "no color"),
            "B": ("subsys name",   f"pre-filled, blue bg {config.COLOR_PREFILLED_BG}"),
            "C": ("FE (frontend designer)", f"pre-filled, blue bg {config.COLOR_PREFILLED_BG}"),
            "D": ("BC (backend coordinator)", f"pre-filled, blue bg {config.COLOR_PREFILLED_BG}"),
            "E": ("BE (backend designer)", f"pre-filled, blue bg {config.COLOR_PREFILLED_BG}"),
            "F": ("ppt upload status",  f"user-filled, yellow bg {config.COLOR_USERFILL_BG}"),
            "G": ("netlist upload status", f"user-filled, yellow bg {config.COLOR_USERFILL_BG}"),
            "H": ("sdc upload status",  f"user-filled, yellow bg {config.COLOR_USERFILL_BG}"),
            "I": ("ccf upload status",  f"user-filled, yellow bg {config.COLOR_USERFILL_BG}"),
            "J": ("upf upload status",  f"user-filled, yellow bg {config.COLOR_USERFILL_BG}"),
            "K": ("PD DEF release date", "green bg FF00B050"),
            "L": ("Remarks", "dark-blue bg FF002060"),
        }
        for col_letter, (role, note) in col_role.items():
            lines.append(f"  {col_letter}: {role}  [{note}]")
        lines.append("")

        # Fill legend
        lines.append("── Fill Legend (G:J columns) ───────────────────────────────────────")
        lines.append("  'v'    → file uploaded to exchange/perforce")
        lines.append("  'x'    → not required for this subsys")
        lines.append("  blank  → not yet uploaded")
        lines.append("  'done' or 'eta:YYYY-MM-DD' for F (ppt) column")
        lines.append("")

        # Skip heuristic
        lines.append("── Row Skip Heuristic ──────────────────────────────────────────────")
        lines.append("  If the font color of column B (subsys) is NOT black (FF000000 or")
        lines.append("  default 00000000), the row is flagged skip=True and excluded from")
        lines.append("  processing.  This catches grey-font rows (e.g. out-of-scope subsys).")
        lines.append("")

        # Deadline context
        lines.append("── Deadline Information ────────────────────────────────────────────")
        deadline_f  = self._cell_value(1, config.COL_PPT)
        deadline_gj = self._cell_value(1, config.COL_NETLIST)   # merged G1:J1
        lines.append(f"  F (ppt) deadline  : {deadline_f!r}")
        lines.append(f"  G:J (netlist…upf) : {deadline_gj!r}")
        lines.append("")

        # Owner notes
        lines.append("── Program Owner Notes ─────────────────────────────────────────────")
        if owner_notes.strip():
            for note_line in owner_notes.strip().splitlines():
                lines.append(f"  {note_line}")
        else:
            lines.append("  (no owner notes provided — add your comments here)")
        lines.append("")
        lines.append("=" * 70)
        lines.append("END OF SHEET DESCRIPTION")
        lines.append("=" * 70)

        return "\n".join(lines)

    def save_llm_prompt(self, filepath: str, owner_notes: str = ""):
        """Write the LLM prompt to a text file for reference."""
        prompt = self.generate_llm_prompt(owner_notes)
        with open(filepath, "w", encoding="utf-8") as fh:
            fh.write(prompt)
        logger.info(f"[ExcelReader] LLM prompt saved to {filepath}")
        return prompt


# ─────────────────────────────────────────────────────────────
# Colour label helper
# ─────────────────────────────────────────────────────────────

_COLOR_LABELS = {
    "FF00B0F0": "blue-prefilled",
    "FFFFC000": "yellow-userfill",
    "FF002060": "dark-blue",
    "FF00B050": "green",
    "FF000000": "black",
    "00000000": "default/transparent",
}

def _color_name(rgb: str) -> str:
    return _COLOR_LABELS.get(rgb, rgb)
